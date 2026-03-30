"""
Parser pasportového Excelu (DPU Energy šablona).

Extrahuje kompletní data:
  - Obecné informace: název, adresa, rok výstavby, rekonstrukce, stav obálky,
                      stav rozvodů, zdroj tepla, osvětlení, chlazení, kuchyň
  - Legenda svítidel: výkon a druh dle označení
  - Přehled (místnosti): osvětlení, OT, TUV, chlazení, větrání
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Union

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ─────────────────────────────────────────────────────────────────────────────
# Pomocné funkce
# ─────────────────────────────────────────────────────────────────────────────

def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.lower() in ("true", "1", "ano", "yes")
    return bool(v)


def _int(v, default: int = 0) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _float(v, default: float = 0.0) -> float:
    if isinstance(v, str):
        v = v.replace(",", ".")
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _stav_z_checkboxu(problem_flags: list) -> str:
    """Bool přiznaky → kategorie stavu (negativní indikátory = problémy)."""
    count = sum(1 for f in problem_flags if f is True)
    if count >= 2:
        return "Vyžaduje rekonstrukci"
    if count == 1:
        return "Uspokojivý"
    return "Dobrý"


def _map_druh(druh: str) -> str:
    d = druh.lower()
    if "led" in d:
        return "LED"
    if "t 5" in d or "t5" in d:
        return "Lineární fluorescenční (T8/T5)"
    if "zářivk" in d or "fluoresc" in d:
        return "Lineární fluorescenční (T8/T5)"
    if "úsporná" in d:
        return "Kompaktní fluorescenční (CFL)"
    if "hps" in d or "výbojk" in d or "halog" in d or "sodík" in d or "rtuť" in d:
        return "Halogenové / výbojkové"
    if "žárovk" in d:
        return "Halogenové / výbojkové"
    return "Jiný"


def _hours_from_room(name: str) -> int:
    n = name.lower()
    if any(w in n for w in ["třída", "učebna", "english", "výtvarná", "fyzika",
                             "chemie", "příroda", "počítač", "jazyková", "hudební"]):
        return 1_800
    if any(w in n for w in ["kabinet", "kancelář", "vedení", "ředitel", "zástupce",
                             "sekretariát", "sborovna"]):
        return 2_500
    if any(w in n for w in ["chodba", "vestibul", "schodiště", "schody", "vstup",
                             "hala", "foyer", "spojovací"]):
        return 3_500
    if any(w in n for w in ["wc", "toaleta", "záchod", "hygien", "umývárna",
                             "sprcha", "šatna"]):
        return 2_000
    if any(w in n for w in ["tělocvična", "sportovní", "gym", "hřiště"]):
        return 1_500
    if any(w in n for w in ["sklad", "technická", "kotelna", "strojovna",
                             "archiv", "dílna", "komora", "úklidová"]):
        return 1_000
    if any(w in n for w in ["jídelna", "kuchyň", "výdejna", "restaurace", "kantýna"]):
        return 2_500
    if any(w in n for w in ["knihovna", "čítárna", "studovna", "aula", "konferenc"]):
        return 2_000
    return 2_000


# ─────────────────────────────────────────────────────────────────────────────
# Výstupní datová třída
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class PasportData:
    # Identifikace
    nazev_objektu: str = ""
    adresa: str = ""
    rok_vystavby: int = 0
    rok_penb: int = 0
    # Stav obálky budovy
    stav_steny: str = "Uspokojivý"
    stav_okna: str = "Uspokojivý"
    stav_strecha: str = "Uspokojivý"
    # Stav techniky
    stav_elektro: str = "Uspokojivý"
    stav_voda: str = "Uspokojivý"
    stav_rozvody_tepla: str = "Uspokojivý"
    stav_ot: str = "Uspokojivý"
    # Rozvody vody
    vod_material_sv: str = "Ocelové"
    vod_material_tv: str = "Ocelové"
    vod_cirkulace: bool = False
    # Vytápění
    vyt_zdroje: list = field(default_factory=list)    # sys_vyt_zdroje
    vyt_vetvi: list = field(default_factory=list)     # sys_vyt_vetvi
    vyt_regulace: list = field(default_factory=list)  # sys_vyt_regulace
    # TUV
    tuv_zdroje: list = field(default_factory=list)    # sys_tuv_zdroje
    tuv_cirkulace: bool = False
    # Chlazení
    chl_instalovano: bool = False
    chl_jednotky: list = field(default_factory=list)  # sys_chl_jednotky
    # VZT
    vzt_jednotky: list = field(default_factory=list)  # sys_vzt_jednotky
    # Osvětlení
    osv_zony: list = field(default_factory=list)      # sys_osv_zony
    # Rekonstrukce
    rekonstrukce: list = field(default_factory=list)  # [{rok, predmet, dokumentace}]
    # Popisné texty (pro Word export)
    popis_stavba: str = ""
    popis_vytapeni: str = ""
    popis_tuv: str = ""
    popis_chlazeni: str = ""
    popis_vzt: str = ""
    popis_osvetleni: str = ""
    popis_elektro: str = ""
    popis_voda_rozv: str = ""
    # Varování parseru
    warnings: list = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Hlavní parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_pasport_xlsx(source: Union[str, Path, io.IOBase]) -> PasportData:
    """
    Parsuje pasportový Excel DPU Energy šablony.
    source: cesta k souboru nebo BytesIO (Streamlit uploader).
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl není nainstalován. Spusťte: pip install openpyxl")

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    d = PasportData()

    # ── Obecné informace ──────────────────────────────────────────────────────
    _parse_obecne(wb, d)

    # ── Legenda svítidel ──────────────────────────────────────────────────────
    legenda = _parse_legenda(wb, d)

    # ── Přehled (místnosti) ───────────────────────────────────────────────────
    _parse_prehled(wb, d, legenda)

    # ── Popisné texty ─────────────────────────────────────────────────────────
    _build_popis(d)

    wb.close()
    return d


# ─────────────────────────────────────────────────────────────────────────────
# Pomocné parsovací funkce
# ─────────────────────────────────────────────────────────────────────────────

def _parse_obecne(wb, d: PasportData) -> None:
    if "Obecné informace" not in wb.sheetnames:
        d.warnings.append("List 'Obecné informace' nenalezen.")
        return

    rows = list(wb["Obecné informace"].iter_rows(values_only=True))

    def row(i):
        return rows[i] if i < len(rows) else (None,) * 10

    # Identifikace
    d.nazev_objektu = _str(row(0)[0])
    d.adresa = ", ".join(filter(None, [_str(row(1)[6]), _str(row(2)[6])]))

    # Rok výstavby – z textu buňky (např. "přístavba z roku 2019")
    rv_text = _str(row(4)[6])
    m = re.search(r"\b(1[89]\d{2}|20[012]\d)\b", rv_text)
    if m:
        d.rok_vystavby = int(m.group())
    else:
        # zkus najít číslo přímo
        for cell in row(4):
            try:
                v = int(cell)
                if 1800 <= v <= 2035:
                    d.rok_vystavby = v
                    break
            except (TypeError, ValueError):
                pass

    # PENB
    try:
        d.rok_penb = _int(row(5)[9])
    except Exception:
        pass

    # Rekonstrukce (řádky 8–12, col 0 = rok, col 3 = předmět, col 7 = dokumentace)
    for ri in range(8, 13):
        r = row(ri)
        try:
            rok = int(r[0])
        except (TypeError, ValueError):
            continue
        if not (1800 <= rok <= 2035):
            continue
        predmet = _str(r[3])
        dok = _str(r[7])
        if predmet and predmet not in ("Předmět rekonstrukce", "vyplnit stupeň dokumentace"):
            d.rekonstrukce.append({"rok": rok, "predmet": predmet, "dokumentace": dok})

    # Stav obvodového pláště (řádky 14–19, col 0 = bool příznak problému)
    fasada_flags = [_bool(row(ri)[0]) for ri in range(14, 20) if row(ri)[0] is not None]
    d.stav_steny = _stav_z_checkboxu(fasada_flags)

    # Poznámka fasáda (řádek 21, col 0)
    pozn_fasada = _str(row(21)[0])
    if pozn_fasada and "Poznámka" not in pozn_fasada:
        d.popis_stavba += f"Fasáda: {pozn_fasada}. "

    # Stav otvorových výplní (řádky 14–17, col 6)
    okna_flags = [_bool(row(ri)[6]) for ri in range(14, 18) if row(ri)[6] is not None and isinstance(row(ri)[6], bool)]
    d.stav_okna = _stav_z_checkboxu(okna_flags)
    pozn_okna = _str(row(19)[6])
    if pozn_okna and "Poznámka" not in pozn_okna:
        d.popis_stavba += f"Okna: {pozn_okna}. "

    # Stav střechy (řádky 24–27, col 0)
    strecha_flags = [_bool(row(ri)[0]) for ri in range(24, 28) if row(ri)[0] is not None and isinstance(row(ri)[0], bool)]
    d.stav_strecha = _stav_z_checkboxu(strecha_flags)
    pozn_strecha = _str(row(29)[0])
    if pozn_strecha and "Poznámka" not in pozn_strecha:
        d.popis_stavba += f"Střecha: {pozn_strecha}. "

    # Stav rozvodů elektřiny (řádky 32–36, col 0) – True = příznak
    ele_flags = [_bool(row(ri)[0]) for ri in range(32, 37) if row(ri)[0] is not None and isinstance(row(ri)[0], bool)]
    d.stav_elektro = _stav_z_checkboxu(ele_flags)
    # Materiál elektrorozvodů
    ele_labels = {33: "hliník", 34: "měď", 35: "kombinace hliník+měď"}
    ele_mat = [lbl for ri, lbl in ele_labels.items() if _bool(row(ri)[0])]
    if ele_mat:
        d.popis_elektro = f"Rozvody: {', '.join(ele_mat)}. "

    # Stav rozvodů vody (řádky 32–36, col 6)
    vod_flags = [_bool(row(ri)[6]) for ri in range(32, 37) if row(ri)[6] is not None and isinstance(row(ri)[6], bool)]
    d.stav_voda = _stav_z_checkboxu(vod_flags)
    vod_mat_map = {33: ("Ocelové", "Ocelové"), 34: ("Měděné", "Měděné"), 35: ("Plastové (PE/PPR)", "Plastové (PE/PPR)")}
    for ri, (sv, tv) in vod_mat_map.items():
        if _bool(row(ri)[6]):
            d.vod_material_sv = sv
            d.vod_material_tv = tv
            break

    # Stav rozvodů tepla (řádky 45–49, col 0)
    teplo_flags = [_bool(row(ri)[0]) for ri in range(45, 50) if row(ri)[0] is not None and isinstance(row(ri)[0], bool)]
    d.stav_rozvody_tepla = _stav_z_checkboxu(teplo_flags)
    pozn_teplo = _str(row(51)[0])
    if pozn_teplo and "Poznámka" not in pozn_teplo:
        d.popis_vytapeni += f"Rozvody: {pozn_teplo}. "

    # Stav OT (řádky 45–49, col 6)
    ot_flags = [_bool(row(ri)[6]) for ri in range(45, 50) if row(ri)[6] is not None and isinstance(row(ri)[6], bool)]
    d.stav_ot = _stav_z_checkboxu(ot_flags)
    pozn_ot = _str(row(51)[6])
    if pozn_ot and "Poznámka" not in pozn_ot:
        d.popis_vytapeni += f"OT: {pozn_ot}. "

    # Zdroj tepla (řádky 54–58, col 6 = bool; col 3 = popis)
    _ZDROJ_TEPLA = {
        54: ("plynová kotelna",    "Plynový kotel",                       0.0, 1, 2000),
        55: ("výměníková stanice", "CZT / horkovod",                      0.0, 1, 2000),
        56: ("tepelné čerpadlo",   "Tepelné čerpadlo vzduch/voda",        0.0, 1, 2015),
        57: ("kogenerační jed.",   "Jiný",                                0.0, 1, 2000),
        58: ("jiný zdroj",         "Jiný",                                0.0, 1, 2000),
    }
    for ri, (label, typ, vykon, pocet, rok) in _ZDROJ_TEPLA.items():
        if _bool(row(ri)[6]):
            d.vyt_zdroje.append({
                "typ": typ, "vykon_kw": vykon,
                "pocet": pocet, "rok": rok, "stav": "Uspokojivý",
            })

    # Vlastnictví / provoz zdroje → poznámka vytápění
    vlastnictvi = _str(row(59)[9])
    provoz = _str(row(60)[9])
    if vlastnictvi:
        d.popis_vytapeni += f"Vlastnictví zdroje: {vlastnictvi}. "
    if provoz:
        d.popis_vytapeni += f"Provoz: {provoz}. "
    pozn_zdroj = _str(row(62)[6])
    if pozn_zdroj and "pohledový stav" not in pozn_zdroj:
        d.popis_vytapeni += pozn_zdroj + ". "

    # Chlazení (řádek 65–66, col 6)
    if _bool(row(65)[6]):   # centrální zdroj chladu
        d.chl_instalovano = True
        d.chl_jednotky.append({
            "typ": "Chiller + fancoily", "vykon_kw": 0.0,
            "rok": 2000, "stav": "Uspokojivý",
        })
    if _bool(row(66)[6]):   # lokální zdroje chladu
        d.chl_instalovano = True
        d.chl_jednotky.append({
            "typ": "Klimatizace split / multi-split", "vykon_kw": 0.0,
            "rok": 2000, "stav": "Uspokojivý",
        })
    pozn_chl = _str(row(71)[6])
    if pozn_chl and "Poznámka" not in pozn_chl:
        d.popis_chlazeni = pozn_chl + "."

    # Centrální VZT (řádek 69, col 6)
    if _bool(row(69)[6]):
        d.vzt_jednotky.append({
            "nazev": "Centrální VZT", "prut_m3h": 0.0,
            "rok": 2000, "zzt": False, "zzt_ucinnost_pct": 0.0,
            "stav": "Uspokojivý",
        })

    # Obecné poznámky
    pozn_obecne = _str(row(74)[0])
    if pozn_obecne:
        # Přiřaď do příslušných popisů
        if "vytápění" in pozn_obecne.lower() or "zóno" in pozn_obecne.lower():
            d.popis_vytapeni += pozn_obecne
        elif "vzducho" in pozn_obecne.lower() or "vzt" in pozn_obecne.lower():
            d.popis_vzt += pozn_obecne
        d.popis_stavba += pozn_obecne


def _parse_legenda(wb, d: PasportData) -> dict:
    legenda: dict = {}
    if "Legenda svítidel" not in wb.sheetnames:
        d.warnings.append("List 'Legenda svítidel' nenalezen – výkony nelze určit.")
        return legenda

    for row in wb["Legenda svítidel"].iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        desig = _str(row[0])
        if not desig or _str(row[1]).startswith("Označení"):
            continue
        druh = _str(row[2])
        pocet = _int(row[3], 1)
        vykon = _float(row[4])
        legenda[desig] = {
            "druh": druh,
            "pocet_zdroju": pocet,
            "vykon_w": vykon,
            "typ_ov": _map_druh(druh),
        }
    return legenda


def _parse_prehled(wb, d: PasportData, legenda: dict) -> None:
    if "Přehled" not in wb.sheetnames:
        d.warnings.append("List 'Přehled' nenalezen.")
        return

    all_rows = list(wb["Přehled"].iter_rows(values_only=True))

    # Najdi hlavičkový řádek
    header_idx = None
    for i, row in enumerate(all_rows):
        if any("NÁZEV MÍSTNOSTI" in _str(c) for c in row):
            header_idx = i
            break
    if header_idx is None:
        d.warnings.append("Hlavička 'NÁZEV MÍSTNOSTI' nenalezena.")
        return

    header = all_rows[header_idx]
    data_start = header_idx + 2

    def _col(needle: str) -> int:
        for i, c in enumerate(header):
            if needle in _str(c).upper():
                return i
        return -1

    C_PODLAZI    = _col("PODLA")   ; C_PODLAZI    = C_PODLAZI    if C_PODLAZI    >= 0 else 6
    C_NAZEV      = _col("NÁZEV MÍSTNOSTI"); C_NAZEV = C_NAZEV if C_NAZEV >= 0 else 8
    C_SVITIDLO   = _col("SVÍTIDLO"); C_SVITIDLO   = C_SVITIDLO   if C_SVITIDLO   >= 0 else 12
    C_POCET_SVIT = _col("POČET SVÍTIDEL"); C_POCET_SVIT = C_POCET_SVIT if C_POCET_SVIT >= 0 else 14
    C_TYP_ZDRO   = _col("TYP ZDROJE")   ; C_TYP_ZDRO   = C_TYP_ZDRO   if C_TYP_ZDRO   >= 0 else 15
    C_POCET_ZDR  = _col("POČET ZDROJŮ") ; C_POCET_ZDR  = C_POCET_ZDR  if C_POCET_ZDR  >= 0 else 18
    C_VYKON_ZDR  = _col("VÝKON ZDROJE") ; C_VYKON_ZDR  = C_VYKON_ZDR  if C_VYKON_ZDR  >= 0 else 19
    C_TYP_OT     = _col("TYP OT")       ; C_TYP_OT     = C_TYP_OT     if C_TYP_OT     >= 0 else 20
    C_POCET_OT   = _col("POČET OT")     ; C_POCET_OT   = C_POCET_OT   if C_POCET_OT   >= 0 else 21
    C_TRV        = _col("TRV")          ; C_TRV        = C_TRV        if C_TRV        >= 0 else 26
    C_TERMO      = _col("TERMO")        ; C_TERMO      = C_TERMO      if C_TERMO      >= 0 else 27
    C_TUV_V_MIST = 34
    C_TUV_ZPUSOB = 35
    C_TUV_ZAR    = 36
    C_TUV_VYKON  = 37
    C_TUV_OBJEM  = 38
    C_CHL_V_MIST = 39
    C_CHL_ZPUSOB = 40
    C_CHL_VYKON  = 41
    C_VZT_FORMA  = 42
    C_VZT_ZPUSOB = 43
    C_VZT_VYKON  = 44

    # Akumulační slovníky
    osv_acc:  dict = {}   # (podlazi, typ_ov) → {prikon_w, pocet, hod_w}
    ot_acc:   dict = {}   # (podlazi, typ_ot) → {pocet_ot, trv_count, total_count, irc_count}
    tuv_cnt:  dict = {}   # zpusob (centrální/lokální) → {pocet_mist, max_vykon, max_objem, typ_zar}
    vzt_acc:  dict = {}   # (podlazi, forma) → {vykon_sum, pocet}
    chl_acc:  dict = {}   # zpusob → {vykon_sum, pocet}

    for row in all_rows[data_start:]:
        if len(row) <= C_SVITIDLO:
            continue
        podlazi = _str(row[C_PODLAZI])
        nazev   = _str(row[C_NAZEV])
        if not podlazi or podlazi in ("-", "None"):
            continue

        # ── Osvětlení ──
        svitidlo = _str(row[C_SVITIDLO])
        if svitidlo and svitidlo not in ("-", "None"):
            pocet_svit = _int(row[C_POCET_SVIT])
            if pocet_svit > 0:
                if svitidlo.upper() == "LED":
                    typ_ov = "LED"
                    vykon_svit_w = 0.0
                elif svitidlo in legenda:
                    leg = legenda[svitidlo]
                    typ_ov = leg["typ_ov"]
                    vykon_svit_w = leg["pocet_zdroju"] * leg["vykon_w"]
                else:
                    pz = _int(row[C_POCET_ZDR], 1)
                    vz = _float(row[C_VYKON_ZDR])
                    typ_zdro = _str(row[C_TYP_ZDRO])
                    typ_ov = _map_druh(typ_zdro) if typ_zdro and typ_zdro != "-" else "Jiný"
                    vykon_svit_w = pz * vz
                hodiny = _hours_from_room(nazev)
                prikon_w = pocet_svit * vykon_svit_w
                key = (podlazi, typ_ov)
                if key not in osv_acc:
                    osv_acc[key] = {"prikon_w": 0.0, "pocet": 0, "hod_w": 0.0}
                osv_acc[key]["pocet"]    += pocet_svit
                osv_acc[key]["prikon_w"] += prikon_w
                osv_acc[key]["hod_w"]    += hodiny * prikon_w

        # ── Otopná tělesa ──
        typ_ot = _str(row[C_TYP_OT])
        if typ_ot and typ_ot not in ("-", "None"):
            pocet_ot = _int(row[C_POCET_OT])
            trv = _str(row[C_TRV]).lower()
            termo = _str(row[C_TERMO]).lower()
            key = (podlazi, typ_ot)
            if key not in ot_acc:
                ot_acc[key] = {"pocet_ot": 0, "trv_count": 0, "irc_count": 0, "total": 0}
            ot_acc[key]["pocet_ot"] += pocet_ot
            ot_acc[key]["total"]    += 1
            if "ano" in trv:
                ot_acc[key]["trv_count"] += 1
            if "irc" in termo:
                ot_acc[key]["irc_count"] += 1

        # ── TUV ──
        if len(row) > C_TUV_V_MIST:
            tuv_v_mist = _str(row[C_TUV_V_MIST]).lower()
            if "ano" in tuv_v_mist:
                zpusob = _str(row[C_TUV_ZPUSOB]).lower() if len(row) > C_TUV_ZPUSOB else "centrální"
                vykon_t = _float(row[C_TUV_VYKON]) if len(row) > C_TUV_VYKON else 0.0
                objem_t = _float(row[C_TUV_OBJEM]) if len(row) > C_TUV_OBJEM else 0.0
                zar     = _str(row[C_TUV_ZAR])     if len(row) > C_TUV_ZAR    else ""
                k = zpusob if zpusob in ("centrální", "lokální") else "centrální"
                if k not in tuv_cnt:
                    tuv_cnt[k] = {"pocet": 0, "max_vykon": 0.0, "max_objem": 0.0, "typ_zar": zar}
                tuv_cnt[k]["pocet"] += 1
                tuv_cnt[k]["max_vykon"] = max(tuv_cnt[k]["max_vykon"], vykon_t)
                tuv_cnt[k]["max_objem"] = max(tuv_cnt[k]["max_objem"], objem_t)

        # ── Chlazení ──
        if len(row) > C_CHL_V_MIST:
            chl_v_mist = _str(row[C_CHL_V_MIST]).lower()
            if "ano" in chl_v_mist:
                d.chl_instalovano = True
                zpusob_c = _str(row[C_CHL_ZPUSOB]).lower() if len(row) > C_CHL_ZPUSOB else "lokální"
                vykon_c  = _float(row[C_CHL_VYKON]) if len(row) > C_CHL_VYKON else 0.0
                k = zpusob_c if zpusob_c in ("centrální", "lokální") else "lokální"
                if k not in chl_acc:
                    chl_acc[k] = {"vykon_sum": 0.0, "pocet": 0}
                chl_acc[k]["vykon_sum"] += vykon_c
                chl_acc[k]["pocet"]     += 1

        # ── Větrání ──
        if len(row) > C_VZT_FORMA:
            forma = _str(row[C_VZT_FORMA]).lower()
            if "nucen" in forma:
                zpusob_v = _str(row[C_VZT_ZPUSOB]).lower() if len(row) > C_VZT_ZPUSOB else ""
                vykon_v  = _float(row[C_VZT_VYKON]) if len(row) > C_VZT_VYKON else 0.0
                k = (podlazi, zpusob_v)
                if k not in vzt_acc:
                    vzt_acc[k] = {"vykon_sum": 0.0, "pocet": 0}
                vzt_acc[k]["vykon_sum"] += vykon_v
                vzt_acc[k]["pocet"]     += 1

    # ── Převod akumulátorů → session-state listy ──────────────────────────────

    # Osvětlení → sys_osv_zony
    for (podlazi, typ_ov), a in sorted(osv_acc.items()):
        prikon_kw = round(a["prikon_w"] / 1000, 2)
        if a["prikon_w"] > 0:
            hod_avg = a["hod_w"] / a["prikon_w"]
        else:
            hod_avg = 2_000
        hodiny = int(round(hod_avg / 100) * 100)
        hodiny = max(500, min(8_760, hodiny))
        d.osv_zony.append({
            "nazev":      f"{podlazi} – {typ_ov}",
            "typ":        typ_ov,
            "prikon_kw":  prikon_kw,
            "pocet":      a["pocet"],
            "rok":        2_000,
            "hodiny_rok": hodiny,
            "rizeni":     "Ruční spínání",
            "stav":       "Uspokojivý",
        })

    # OT → sys_vyt_vetvi
    _OT_TYP_MAP = {
        "deskov":     "Dvoutrubková",
        "litinovám":  "Dvoutrubková",
        "litinov":    "Dvoutrubková",
        "podlahov":   "Podlahové vytápění",
        "teplovzduš": "Dvoutrubková",
    }
    for (podlazi, typ_ot), a in sorted(ot_acc.items()):
        typ_soustava = "Dvoutrubková"
        for k, v in _OT_TYP_MAP.items():
            if k in typ_ot.lower():
                typ_soustava = v
                break
        has_trv = a["trv_count"] > a["total"] * 0.5
        has_irc = a["irc_count"] > 0
        d.vyt_vetvi.append({
            "typ":      typ_soustava,
            "popis":    f"{podlazi} – {typ_ot}",
            "pocet_ot": a["pocet_ot"],
            "trv":      has_trv or has_irc,
            "rok":      2_000,
            "stav":     "Uspokojivý",
        })

    # Regulace – odvození z TRV/IRC dat
    if d.vyt_vetvi:
        has_any_irc = any(
            a["irc_count"] > 0 for a in ot_acc.values()
        )
        has_any_trv = any(
            a["trv_count"] > a["total"] * 0.5 for a in ot_acc.values()
        )
        if has_any_irc:
            reg_typ = "Ekvitermní + TRV + individuální regulace (IRC)"
        elif has_any_trv:
            reg_typ = "Ekvitermní + termostatické ventily (TRV)"
        else:
            reg_typ = "Ekvitermní regulace kotelny"
        d.vyt_regulace.append({"typ": reg_typ, "rok": 2000, "stav": "Uspokojivý"})

    # TUV → sys_tuv_zdroje
    _TUV_TYP_MAP = {
        "centrální": "Plynový zásobníkový ohřívač",
        "lokální":   "Elektrický bojler",
    }
    for zpusob, t in tuv_cnt.items():
        typ_tuv = _TUV_TYP_MAP.get(zpusob, "Jiný")
        # Upřesni typ dle zdroje tepla z Obecné informace
        if d.vyt_zdroje:
            zdroj_typ = d.vyt_zdroje[0]["typ"]
            if "CZT" in zdroj_typ:
                typ_tuv = "Ohřev z CZT"
            elif "Tepelné čerpadlo" in zdroj_typ:
                typ_tuv = "Tepelné čerpadlo (vzduch/voda)"
        d.tuv_zdroje.append({
            "typ":     typ_tuv,
            "objem_l": t["max_objem"] if t["max_objem"] > 0 else 0.0,
            "rok":     2_000,
            "stav":    "Uspokojivý",
        })

    # Chlazení z Přehledu → doplnit/rozšířit sys_chl_jednotky
    _CHL_TYP_MAP = {"centrální": "Chiller + fancoily", "lokální": "Klimatizace split / multi-split"}
    for zpusob_c, a in chl_acc.items():
        # Přidej jen pokud ještě není z Obecné informace
        already = any(
            (_CHL_TYP_MAP.get(zpusob_c, "") in j.get("typ", ""))
            for j in d.chl_jednotky
        )
        if not already:
            d.chl_jednotky.append({
                "typ":      _CHL_TYP_MAP.get(zpusob_c, "Klimatizace split / multi-split"),
                "vykon_kw": round(a["vykon_sum"], 1),
                "rok":      2_000,
                "stav":     "Uspokojivý",
            })

    # VZT → sys_vzt_jednotky (z Přehledu) – doplní centrální z Obecné informace
    for (podlazi_v, zpusob_v), a in sorted(vzt_acc.items()):
        popis = f"{podlazi_v} – {zpusob_v}" if zpusob_v else podlazi_v
        d.vzt_jednotky.append({
            "nazev":            popis,
            "prut_m3h":         0.0,
            "rok":              2_000,
            "zzt":              False,
            "zzt_ucinnost_pct": 0.0,
            "stav":             "Uspokojivý",
        })

    if not d.osv_zony:
        d.warnings.append("Nebyla nalezena žádná svítidla v listu 'Přehled'.")


def _build_popis(d: PasportData) -> None:
    """Doplní a uzavře popisné texty na základě strukturovaných dat."""

    # Stavba
    if d.rok_vystavby:
        d.popis_stavba = f"Rok výstavby: {d.rok_vystavby}. " + d.popis_stavba
    if d.rekonstrukce:
        rek_list = "; ".join(f"{r['rok']} – {r['predmet']}" for r in d.rekonstrukce)
        d.popis_stavba += f"Rekonstrukce: {rek_list}. "
    d.popis_stavba = d.popis_stavba.strip()

    # Vytápění
    if d.vyt_zdroje:
        typy = ", ".join(z["typ"] for z in d.vyt_zdroje)
        d.popis_vytapeni = f"Zdroj tepla: {typy}. " + d.popis_vytapeni
    if d.vyt_vetvi:
        ot_prehled = "; ".join(
            f"{v['popis']} ({v['pocet_ot']} ks{', TRV' if v['trv'] else ''})"
            for v in d.vyt_vetvi
        )
        d.popis_vytapeni += f"Otopná tělesa: {ot_prehled}. "
    d.popis_vytapeni = d.popis_vytapeni.strip()

    # TUV
    if d.tuv_zdroje:
        typy_tuv = ", ".join(z["typ"] for z in d.tuv_zdroje)
        d.popis_tuv = f"Příprava TUV: {typy_tuv}. " + d.popis_tuv
    d.popis_tuv = d.popis_tuv.strip()

    # Chlazení
    if not d.chl_instalovano:
        d.popis_chlazeni = "Objekt není vybaven centrálním chladicím systémem. " + d.popis_chlazeni
    d.popis_chlazeni = d.popis_chlazeni.strip()

    # VZT
    if d.vzt_jednotky:
        vzt_prehled = "; ".join(j["nazev"] for j in d.vzt_jednotky if j["nazev"])
        d.popis_vzt = f"Větrací systémy: {vzt_prehled}." if vzt_prehled else d.popis_vzt
    d.popis_vzt = d.popis_vzt.strip()

    # Osvětlení
    if d.osv_zony:
        non_led = [z for z in d.osv_zony if z["typ"] != "LED"]
        led = [z for z in d.osv_zony if z["typ"] == "LED"]
        parts = []
        if led:
            parts.append(f"LED svítidla: {sum(z['pocet'] for z in led)} ks")
        if non_led:
            non_led_prehled = ", ".join(
                f"{z['typ']} {z['pocet']} ks ({z['prikon_kw']:.1f} kW)"
                for z in non_led if z["pocet"] > 0
            )
            if non_led_prehled:
                parts.append(f"non-LED: {non_led_prehled}")
        d.popis_osvetleni = ". ".join(parts) + "." if parts else ""

    # Elektro
    if not d.popis_elektro:
        d.popis_elektro = "Elektroinstalace dle pasportu."

    # Rozvody vody
    if not d.popis_voda_rozv:
        d.popis_voda_rozv = (
            f"Studená voda: {d.vod_material_sv}. "
            f"Teplá voda / cirkulace: {d.vod_material_tv}."
        )
