"""
Generátor HPR – agregovaný slepý výkaz výměr v Excelu.

Výstup: .xlsx se strukturou:
  Krycí list – metadata projektu + souhrnná tabulka
  <OP ID>    – jeden list na opatření s položkami
"""
from __future__ import annotations

from io import BytesIO
from typing import NamedTuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from epc_engine.models import BuildingInfo, ProjectResult
from epc_engine.op_descriptions import OP_INFO, OP_ROZPOCET_POLOZKY

# ── Styling ───────────────────────────────────────────────────────────────────

_FN = "Calibri"

# Barvy (HEX bez #)
_C_DARK   = "282560"   # tmavě modrá – záhlaví sloupců
_C_PURPLE = "5551A6"   # fialová – celkový součet
_C_DPH    = "9FBADF"   # světle modrá – DPH řádky
_C_SECT   = "D5E8F0"   # velmi světle modrá – záhlaví sekce
_C_TOTAL  = "EBF5FB"   # mezisoučty
_C_WHITE  = "FFFFFF"
_C_GRAY   = "808080"

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_INT = "#,##0"
_FMT_DEC = "#,##0.00"

_DPH = 0.21


def _f(bold=False, size=10, color="000000"):
    return Font(name=_FN, bold=bold, size=size, color=color)


def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex6)


def _al(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _cell(ws, row: int, col: int, value=None, *,
          bold=False, size=10, color="000000", bg: str | None = None,
          h="left", v="center", wrap=False, fmt: str | None = None):
    """Zapíše hodnotu a aplikuje styling na buňku."""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = _f(bold=bold, size=size, color=color)
    c.alignment = _al(h=h, v=v, wrap=wrap)
    c.border = _BORDER
    if bg:
        c.fill = _fill(bg)
    if fmt:
        c.number_format = fmt
    return c


def _formula(ws, row: int, col: int, expr: str, *,
             bold=False, size=9, color="000000", bg: str | None = None,
             h="right", fmt=_FMT_INT):
    """Zapíše Excel formuli s pravým zarovnáním a číselným formátem."""
    c = ws.cell(row=row, column=col)
    c.value = expr
    c.font = _f(bold=bold, size=size, color=color)
    c.alignment = _al(h=h)
    c.border = _BORDER
    if bg:
        c.fill = _fill(bg)
    c.number_format = fmt
    return c


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ── Info o listech ──────────────────────────────────────────────────────────

class _SheetRef(NamedTuple):
    name: str       # název listu
    sum_b: int      # řádek se součtem sekce B (G sloupec)
    sum_d: int      # řádek se součtem sekce D (G sloupec)
    total: int      # řádek s celkovým součtem OP (G sloupec)


# ── List jednoho opatření ─────────────────────────────────────────────────────

def _op_sheet(wb, op_idx: int, op_id: str, op_nazev: str,
              investice: float, polozky: list[tuple[str, str]]) -> _SheetRef:
    """Vytvoří list pro jedno OP a vrátí reference na klíčové řádky."""
    ws = wb.create_sheet(title=op_id)

    # Šířky sloupců  A  B   C   D   E   F   G   H   I
    for col, w in zip("ABCDEFGHI", [5, 12, 42, 13, 12, 18, 20, 12, 22]):
        ws.column_dimensions[col].width = w

    # ── Řádek 1: Titulek ────────────────────────────────────────────────────
    _merge(ws, 1, 2, 1, 9)
    _cell(ws, 1, 2, "Agregovaný slepý výkaz výměr – EPC projekt",
          bold=True, size=14, h="center")

    # ── Řádky 2–3: Projekt / objednatel (navázáno na Krycí list) ────────────
    _cell(ws, 2, 2, "Název stavby:", bold=True, size=10)
    _merge(ws, 2, 3, 2, 9)
    ws.cell(row=2, column=3).value = "='Krycí list'!B2"
    ws.cell(row=2, column=3).font = _f(size=10)
    ws.cell(row=2, column=3).border = _BORDER

    _cell(ws, 3, 2, "Objednatel:", bold=True, size=10)
    _merge(ws, 3, 3, 3, 9)
    ws.cell(row=3, column=3).value = "=CONCAT('Krycí list'!D2,\", \",'Krycí list'!D3)"
    ws.cell(row=3, column=3).font = _f(size=10)
    ws.cell(row=3, column=3).border = _BORDER

    # ── Řádky 5–6: Číslo a název opatření ───────────────────────────────────
    _cell(ws, 5, 2, "Číslo opatření:", bold=True, size=10)
    _cell(ws, 5, 3, op_idx, size=10)
    _cell(ws, 6, 2, "Název opatření:", bold=True, size=10)
    _merge(ws, 6, 3, 6, 9)
    _cell(ws, 6, 3, op_nazev, size=10, bold=True)

    # ── Řádek 8: Záhlaví sloupců ────────────────────────────────────────────
    ws.row_dimensions[8].height = 32
    hdrs = [
        ("B", "Číslo položky", "center"),
        ("C", "Popis položky", "left"),
        ("D", "MJ", "center"),
        ("E", "Počet MJ", "center"),
        ("F", "Jednotková cena\nbez DPH [Kč/MJ]", "center"),
        ("G", "Celkem\nbez DPH [Kč]", "center"),
        ("H", "DPH\n21 %", "center"),
        ("I", "Celkem\nvč. DPH [Kč]", "center"),
    ]
    for i, (col_letter, text, align) in enumerate(hdrs):
        col_num = ord(col_letter) - ord("A") + 1
        _cell(ws, 8, col_num, text, bold=True, size=9,
              color=_C_WHITE, bg=_C_DARK, h=align, v="center", wrap=True)

    # ── Řádek 9: Jednotky ───────────────────────────────────────────────────
    for col, text in [(2, "–"), (3, "–"), (4, "–"), (5, "–"),
                      (6, "Kč/MJ"), (7, "Kč"), (8, "21 %"), (9, "Kč")]:
        _cell(ws, 9, col, text, size=9, h="center")

    row = 10

    # ── Sekce B: Stavební práce a dodávky ───────────────────────────────────
    _merge(ws, row, 2, row, 9)
    _cell(ws, row, 2, f"{op_idx}.B   Stavební práce a dodávky",
          bold=True, size=9, bg=_C_SECT)
    row += 1

    b_start = row
    for i, (popis, mj) in enumerate(polozky, start=1):
        _cell(ws, row, 1, i, size=9, h="center")
        _cell(ws, row, 2, f"{op_idx}.B.{i}", size=9)
        _cell(ws, row, 3, popis, size=9, wrap=True)
        ws.row_dimensions[row].height = 15
        _cell(ws, row, 4, mj, size=9, h="center")
        # Množství – prázdné (vyplní uchazeč)
        _cell(ws, row, 5, None, size=9, h="right", fmt=_FMT_DEC)
        # Jedn. cena – prázdné (vyplní uchazeč)
        _cell(ws, row, 6, None, size=9, h="right", fmt=_FMT_INT)
        # Celkem = E * F
        _formula(ws, row, 7, f"=IF(E{row}*F{row}=0,\"\",E{row}*F{row})", size=9, fmt=_FMT_INT)
        # DPH
        _formula(ws, row, 8, f"=IF(G{row}=\"\",\"\",G{row}*{_DPH})", size=9, h="right", fmt=_FMT_INT)
        # Celkem vč. DPH
        _formula(ws, row, 9, f"=IF(G{row}=\"\",\"\",G{row}+H{row})", size=9, h="right", fmt=_FMT_INT)
        row += 1

    b_end = row - 1

    # Mezisoučet B
    sum_b = row
    _cell(ws, row, 1, "Σ", bold=True, size=9, h="center", bg=_C_TOTAL)
    _merge(ws, row, 2, row, 6)
    _cell(ws, row, 2, f"Součet za {op_idx}.B – Stavební práce a dodávky",
          bold=True, size=9, bg=_C_TOTAL)
    _formula(ws, row, 7, f"=SUMIF(G{b_start}:G{b_end},\"<>\")", bold=True, bg=_C_TOTAL)
    _formula(ws, row, 8, f"=G{row}*{_DPH}", bold=True, bg=_C_TOTAL)
    _formula(ws, row, 9, f"=G{row}+H{row}", bold=True, bg=_C_TOTAL)
    row += 2

    # ── Sekce D: Další práce ─────────────────────────────────────────────────
    _merge(ws, row, 2, row, 9)
    _cell(ws, row, 2, f"{op_idx}.D   Další práce a materiál nutný pro dokončení díla",
          bold=True, size=9, bg=_C_SECT)
    row += 1

    d_start = row
    for i in range(1, 4):  # 3 prázdné řádky
        _cell(ws, row, 1, i, size=9, h="center")
        _cell(ws, row, 2, f"{op_idx}.D.{i}", size=9)
        _cell(ws, row, 3, "[může doplnit Dodavatel]", size=9, color=_C_GRAY)
        _cell(ws, row, 4, "komplet", size=9, h="center")
        _cell(ws, row, 5, 1, size=9, h="right")
        _cell(ws, row, 6, None, size=9, h="right", fmt=_FMT_INT)
        _formula(ws, row, 7, f"=IF(F{row}=0,\"\",F{row}*E{row})", size=9, fmt=_FMT_INT)
        _formula(ws, row, 8, f"=IF(G{row}=\"\",\"\",G{row}*{_DPH})", size=9, h="right", fmt=_FMT_INT)
        _formula(ws, row, 9, f"=IF(G{row}=\"\",\"\",G{row}+H{row})", size=9, h="right", fmt=_FMT_INT)
        row += 1

    d_end = row - 1

    # Mezisoučet D
    sum_d = row
    _cell(ws, row, 1, "Σ", bold=True, size=9, h="center", bg=_C_TOTAL)
    _merge(ws, row, 2, row, 6)
    _cell(ws, row, 2, f"Součet za {op_idx}.D – Další práce a materiál",
          bold=True, size=9, bg=_C_TOTAL)
    _formula(ws, row, 7, f"=SUMIF(G{d_start}:G{d_end},\"<>\")", bold=True, bg=_C_TOTAL)
    _formula(ws, row, 8, f"=G{row}*{_DPH}", bold=True, bg=_C_TOTAL)
    _formula(ws, row, 9, f"=G{row}+H{row}", bold=True, bg=_C_TOTAL)
    row += 2

    # ── Celkový součet opatření ──────────────────────────────────────────────
    total_row = row
    _cell(ws, row, 1, "Σ", bold=True, size=10, h="center",
          color=_C_WHITE, bg=_C_PURPLE)
    _merge(ws, row, 2, row, 6)
    _cell(ws, row, 2, f"Celkem za opatření {op_idx} – {op_nazev}",
          bold=True, size=10, color=_C_WHITE, bg=_C_PURPLE)
    _formula(ws, row, 7, f"=G{sum_b}+G{sum_d}",
             bold=True, color=_C_WHITE, bg=_C_PURPLE)
    _formula(ws, row, 8, f"=G{row}*{_DPH}",
             bold=True, color=_C_WHITE, bg=_C_PURPLE)
    _formula(ws, row, 9, f"=G{row}+H{row}",
             bold=True, color=_C_WHITE, bg=_C_PURPLE)
    row += 2

    # ── Orientační poznámka ──────────────────────────────────────────────────
    inv_fmt = f"{investice:,.0f}".replace(",", "\u00a0")
    _merge(ws, row, 1, row, 9)
    c = ws.cell(row=row, column=1,
                value=f"Orientační hodnota dle energetického posudku: {inv_fmt} Kč bez DPH")
    c.font = Font(name=_FN, size=8, italic=True, color=_C_GRAY)
    c.alignment = _al(h="left")

    return _SheetRef(name=op_id, sum_b=sum_b, sum_d=sum_d, total=total_row)


# ── Krycí list ────────────────────────────────────────────────────────────────

def _krytci_list(wb, budova: BuildingInfo, result: ProjectResult,
                 ss: dict, refs: list[tuple[str, _SheetRef]]) -> None:
    ws = wb.create_sheet(title="Krycí list", index=0)

    # Šířky sloupců
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 22

    nazev  = ss.get("zakaz_nazev", "") or (budova.objekt_nazev or "")
    zadav  = ss.get("zadavatel_nazev", "")
    adresa = ss.get("zadavatel_adresa", "")
    ico    = ss.get("zadavatel_ico", "")
    zprac  = ss.get("zpracovatel_zastupce", "")
    datum  = ss.get("datum", "")
    lokal  = ss.get("lokalita_projekt", "") or (budova.objekt_adresa or "")

    # ── Řádek 1: Titulek ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    _merge(ws, 1, 1, 1, 7)
    _cell(ws, 1, 1,
          "Krycí list rozpočtu pro zakázku řešenou metodou Design & Build",
          bold=True, size=16, h="center")

    # ── Řádky 2–3: Název stavby + Objednatel ────────────────────────────────
    _merge(ws, 2, 1, 3, 1)
    _cell(ws, 2, 1, "Název stavby:", bold=True, size=11, v="center")
    _merge(ws, 2, 2, 3, 2)
    _cell(ws, 2, 2, nazev, size=11, wrap=True, v="center")
    _merge(ws, 2, 3, 2, 5)
    _cell(ws, 2, 3, "Objednatel:", bold=True, size=11)
    _merge(ws, 3, 3, 3, 5)
    _cell(ws, 3, 3, zadav, size=11)
    _cell(ws, 2, 6, "IČO:", bold=True, size=11)
    _cell(ws, 2, 7, ico, size=11)

    # ── Řádky 4–5: Druh stavby + adresa ─────────────────────────────────────
    _merge(ws, 4, 1, 5, 1)
    _cell(ws, 4, 1, "Druh stavby:", bold=True, size=11, v="center")
    _merge(ws, 4, 2, 5, 2)
    _cell(ws, 4, 2, "Energeticky úsporná opatření – EPC projekt", size=11, v="center")
    _merge(ws, 4, 3, 5, 5)
    _cell(ws, 4, 3, adresa, size=11, v="center", wrap=True)

    # ── Řádky 6–7: Lokalita + Zpracovatel ────────────────────────────────────
    _merge(ws, 6, 1, 7, 1)
    _cell(ws, 6, 1, "Lokalita:", bold=True, size=11, v="center")
    _merge(ws, 6, 2, 7, 2)
    _cell(ws, 6, 2, lokal, size=11, v="center")
    _merge(ws, 6, 3, 7, 3)
    _cell(ws, 6, 3, "Zpracovatel:", bold=True, size=11, v="center")
    _merge(ws, 6, 4, 7, 7)
    _cell(ws, 6, 4, zprac, size=11, v="center")

    # ── Řádek 8: Datum ───────────────────────────────────────────────────────
    _cell(ws, 8, 1, "Datum:", bold=True, size=11)
    _merge(ws, 8, 2, 8, 7)
    _cell(ws, 8, 2, datum, size=11)

    # ── Řádek 10: Nadpis tabulky ─────────────────────────────────────────────
    _merge(ws, 10, 1, 10, 7)
    _cell(ws, 10, 1, "Rozpočtové náklady v Kč",
          bold=True, size=12, h="center", color=_C_WHITE, bg=_C_DARK)

    # ── Řádek 11: Legenda ────────────────────────────────────────────────────
    _merge(ws, 11, 1, 11, 7)
    _cell(ws, 11, 1,
          "Legenda:   B … Stavební práce a dodávky   |   "
          "D … Další práce a materiál nutný pro dokončení díla",
          size=9)

    # ── Řádek 12: Záhlaví sloupců ────────────────────────────────────────────
    ws.row_dimensions[12].height = 32
    for col, text in enumerate([
        "Č.", "Název opatření",
        "B – Stavební práce\nbez DPH [Kč]",
        "D – Další práce\nbez DPH [Kč]",
        "Celkem bez DPH [Kč]",
        "DPH 21 % [Kč]",
        "Celkem vč. DPH [Kč]",
    ], start=1):
        _cell(ws, 12, col, text, bold=True, size=9,
              color=_C_WHITE, bg=_C_DARK, h="center", v="center", wrap=True)

    # ── Řádky 13+: Jedno OP na řádek ─────────────────────────────────────────
    row = 13
    for op_idx, (op_id, ref) in enumerate(refs, start=1):
        op_nazev = OP_INFO.get(op_id, {}).get("title", op_id)
        _cell(ws, row, 1, op_idx, size=9, h="center")
        _cell(ws, row, 2, f"{op_id} – {op_nazev}", size=9, wrap=True)
        _formula(ws, row, 3, f"='{ref.name}'!G{ref.sum_b}", size=9)
        _formula(ws, row, 4, f"='{ref.name}'!G{ref.sum_d}", size=9)
        _formula(ws, row, 5, f"=C{row}+D{row}", size=9)
        _formula(ws, row, 6, f"=E{row}*{_DPH}", size=9)
        _formula(ws, row, 7, f"=E{row}+F{row}", size=9)
        row += 1

    # ── Řádek celkový součet ──────────────────────────────────────────────────
    first_data = 13
    last_data = row - 1
    _cell(ws, row, 1, "Σ", bold=True, size=10, h="center",
          color=_C_WHITE, bg=_C_PURPLE)
    _cell(ws, row, 2, "Celkové náklady – všechna opatření",
          bold=True, size=10, color=_C_WHITE, bg=_C_PURPLE)
    for col in [3, 4, 5, 6, 7]:
        col_l = get_column_letter(col)
        _formula(ws, row, col,
                 f"=SUM({col_l}{first_data}:{col_l}{last_data})",
                 bold=True, color=_C_WHITE, bg=_C_PURPLE)
    total_row = row
    row += 2

    # ── Základ DPH ───────────────────────────────────────────────────────────
    _cell(ws, row, 1, "Základ pro DPH:", bold=True, size=9, bg=_C_DPH)
    _formula(ws, row, 2, f"=E{total_row}",
             bold=True, size=9, bg=_C_DPH, h="right")
    _merge(ws, row, 3, row, 5)
    _cell(ws, row, 3, "(základ = celkem bez DPH)", size=9, bg=_C_DPH, color=_C_GRAY)
    row += 1

    _cell(ws, row, 1, "Základ DPH 21 %:", bold=True, size=9, bg=_C_DPH)
    _formula(ws, row, 2, f"=B{row-1}", bold=True, size=9, bg=_C_DPH, h="right")
    _cell(ws, row, 3, "DPH 21 %:", bold=True, size=9, bg=_C_DPH)
    _formula(ws, row, 4, f"=B{row}*{_DPH}", bold=True, size=9, bg=_C_DPH)
    _merge(ws, row, 5, row, 6)
    _cell(ws, row, 5, "Celkem včetně DPH:", bold=True, size=10, bg=_C_DPH, h="right")
    _formula(ws, row, 7, f"=B{row}+D{row}",
             bold=True, size=10, bg=_C_DPH)
    row += 2

    # ── Sekce pro podpisy ─────────────────────────────────────────────────────
    _merge(ws, row, 1, row + 3, 2)
    c = ws.cell(row=row, column=1, value="Objednatel")
    c.font = _f(bold=True, size=11)
    c.alignment = _al(h="center", v="center")
    c.border = _BORDER
    _merge(ws, row, 4, row + 3, 7)
    c2 = ws.cell(row=row, column=4, value="Zhotovitel")
    c2.font = _f(bold=True, size=11)
    c2.alignment = _al(h="center", v="center")
    c2.border = _BORDER
    row += 4
    _merge(ws, row, 1, row, 2)
    _cell(ws, row, 1, "Datum, razítko a podpis", size=9, h="center")
    _merge(ws, row, 4, row, 7)
    _cell(ws, row, 4, "Datum, razítko a podpis", size=9, h="center")


# ── Veřejné API ───────────────────────────────────────────────────────────────

def generuj_hpr(
    budova: BuildingInfo,
    result: ProjectResult,
    aktivni_op: list[str],
    ss: dict,
) -> BytesIO:
    """
    Sestaví Excel sešit – agregovaný slepý výkaz výměr pro VŘ EPC.

    Parametry:
        budova      – BuildingInfo ze session state
        result      – ProjectResult z build_project().vypocitej()
        aktivni_op  – seznam aktivních OP ID (např. [\"OP1a\", \"OP7\"])
        ss          – st.session_state jako dict
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # smazat výchozí prázdný list

    # Nejdřív vytvořit listy OP (potřebujeme znát čísla řádků pro Krycí list)
    refs: list[tuple[str, _SheetRef]] = []
    for op_idx, op_id in enumerate(aktivni_op, start=1):
        mr = next((m for m in result.vysledky if m.id == op_id), None)
        investice = mr.investice if mr else 0.0
        op_nazev = OP_INFO.get(op_id, {}).get("title", op_id)
        polozky = OP_ROZPOCET_POLOZKY.get(op_id, [("Dodávka a montáž", "komplet")])
        ref = _op_sheet(wb, op_idx, op_id, op_nazev, investice, polozky)
        refs.append((op_id, ref))

    # Pak Krycí list (index=0 = první pozice)
    _krytci_list(wb, budova, result, ss, refs)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
